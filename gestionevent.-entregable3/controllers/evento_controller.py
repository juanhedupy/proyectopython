import tkinter as tk
from tkinter import messagebox
from datetime import datetime
from views.evento_view import EventoView
from config import COLORES_BOTONES
from tkinter import filedialog  # ← AGREGAR: Para diálogo de guardado
import openpyxl  # ← AGREGAR: pip install openpyxl si no lo tienes
from openpyxl.styles import Font, Alignment, PatternFill  # ← AGREGAR: Para estilos


class EventoController:
    def __init__(self, parent, app, evento_model, cliente_model, recinto_model):
        self.app = app
        self.model = evento_model
        self.cliente_model = cliente_model
        self.recinto_model = recinto_model
        self.view = EventoView(parent, self)
        self.id_actual = None
        self.cargar_recintos()

    def get_color_boton(self, tipo):
        return COLORES_BOTONES.get(tipo, '#gray')

    def cargar_recintos(self):
        recintos = self.recinto_model.obtener_todos()
        if recintos:
            valores_recintos = [f"{r[0]} - {r[1]}" for r in recintos]
            self.view.widgets['id_recinto']['values'] = valores_recintos

    def guardar(self):
        valores = self.view.obtener_valores()
        campos_obligatorios = ['id_recinto', 'identificacion_cliente', 'titulo', 'tipo',
                               'fecha_hora_inicio', 'fecha_hora_fin', 'estado', 'costo']
        for campo in campos_obligatorios:
            if not valores[campo]:
                messagebox.showerror("Error", f"El campo {campo} es obligatorio")
                return

        try:
            # Extraer ID del recinto
            id_recinto = int(valores['id_recinto'].split(' - ')[0])

            # Obtener ID del cliente
            resultado_cliente = self.cliente_model.buscar_por_identificacion(valores['identificacion_cliente'])
            if not resultado_cliente:
                messagebox.showerror("Error", "No se encontró un cliente con esa identificación")
                return
            id_cliente = resultado_cliente[0][0]

            # Convertir numéricos
            asistentes = int(valores['asistentes_estimados']) if valores['asistentes_estimados'] else 0
            costo = float(valores['costo'])

            resultado = self.model.guardar(
                id_recinto, id_cliente, valores['titulo'], valores['tipo'],
                valores['fecha_hora_inicio'], valores['fecha_hora_fin'],
                asistentes, valores['estado'], valores.get('descripcion', ''), costo
            )

            if resultado is not None:
                messagebox.showinfo("Éxito", "Evento guardado correctamente")
                self.view.limpiar_formulario()
                self.app.id_evento_actual = None

        except ValueError as ve:
            messagebox.showerror("Error", f"Asistentes y costo deben ser valores numéricos: {ve}")
        except Exception as e:
            messagebox.showerror("Error", f"Error al guardar evento: {e}")

    def actualizar(self):
        if not self.app.id_evento_actual:
            messagebox.showwarning("Advertencia", "Primero debe buscar un evento para actualizar")
            return

        valores = self.view.obtener_valores()
        campos_obligatorios = ['id_recinto', 'identificacion_cliente', 'titulo', 'tipo',
                               'fecha_hora_inicio', 'fecha_hora_fin', 'estado', 'costo']
        for campo in campos_obligatorios:
            if not valores[campo]:
                messagebox.showerror("Error", f"El campo {campo} es obligatorio")
                return

        try:
            # Extraer ID del recinto
            id_recinto = int(valores['id_recinto'].split(' - ')[0])

            # Obtener ID del cliente
            resultado_cliente = self.cliente_model.buscar_por_identificacion(valores['identificacion_cliente'])
            if not resultado_cliente:
                messagebox.showerror("Error", "No se encontró un cliente con esa identificación")
                return
            id_cliente = resultado_cliente[0][0]

            # Convertir numéricos
            asistentes = int(valores['asistentes_estimados']) if valores['asistentes_estimados'] else 0
            costo = float(valores['costo'])

            resultado = self.model.actualizar(
                self.app.id_evento_actual, id_recinto, id_cliente, valores['titulo'], valores['tipo'],
                valores['fecha_hora_inicio'], valores['fecha_hora_fin'],
                asistentes, valores['estado'], valores.get('descripcion', ''), costo
            )

            messagebox.showinfo("Éxito", "Evento actualizado correctamente")
            self.app.id_evento_actual = None

        except ValueError as ve:
            messagebox.showerror("Error", f"Error en formatos numéricos: {ve}")
        except Exception as e:
            messagebox.showerror("Error", f"Error al actualizar evento: {e}")

    def buscar(self):
        ventana, entry, listbox = self.view.ventana_buscar()

        def realizar_busqueda():
            identificacion = entry.get().strip()
            if not identificacion:
                messagebox.showwarning("Advertencia", "Por favor ingrese una identificación")
                return

            try:
                # Obtener ID cliente
                resultado_cliente = self.cliente_model.buscar_por_identificacion(identificacion)
                if not resultado_cliente:
                    messagebox.showinfo("Información", "No se encontró ningún cliente con esa identificación")
                    return
                id_cliente = resultado_cliente[0][0]

                # Obtener eventos
                resultados = self.model.obtener_por_cliente(id_cliente)
                listbox.delete(0, tk.END)

                if not resultados:
                    messagebox.showinfo("Información", "El cliente no tiene eventos asociados")
                    return

                for resultado in resultados:
                    fecha_inicio = resultado[5].strftime('%Y-%m-%d %H:%M') if isinstance(resultado[5], datetime) else str(resultado[5])
                    listbox.insert(tk.END, f"{resultado[3]} - {fecha_inicio} - {resultado[11]}")  # titulo - fecha - nombre_recinto
                listbox.resultados = resultados

            except Exception as e:
                messagebox.showerror("Error", f"Error al buscar: {e}")

        def seleccionar_evento(event):
            if not listbox.curselection():
                return
            index = listbox.curselection()[0]
            evento = listbox.resultados[index]
            self.mostrar_evento(evento)
            ventana.destroy()

        btn_buscar = tk.Button(ventana, text="Buscar", command=realizar_busqueda, bg='#3498db', fg='white')
        btn_buscar.pack(pady=10)

        entry.bind('<Return>', lambda event: realizar_busqueda())
        listbox.bind('<Double-Button-1>', seleccionar_evento)

    def mostrar_evento(self, evento):
        self.view.mostrar_datos(evento)
        self.app.id_evento_actual = evento[0]

    def eliminar(self):
        if not self.app.id_evento_actual:
            messagebox.showwarning("Advertencia", "Primero debe buscar un evento para eliminar")
            return

        if messagebox.askyesno("Confirmar eliminación", "¿Está seguro de que desea eliminar este evento? Esta acción no se puede deshacer."):
            self.model.eliminar(self.app.id_evento_actual)
            messagebox.showinfo("Éxito", "Evento eliminado correctamente")
            self.view.limpiar_formulario()
            self.app.id_evento_actual = None

    def limpiar(self):
        self.view.limpiar_formulario()
        self.app.id_evento_actual = None

    def exportar_excel(self):
        """
        Exporta todos los eventos a un archivo Excel usando ObtenerTodosLosEventos.
        """
        try:
            print("DEBUG EXPORT: Iniciando exportación...")  # Debug temporal
            eventos = self.model.obtener_todos()  # Llama al método corregido
            print(f"DEBUG EXPORT: Obtenidos {len(eventos)} eventos de BD")  # Debug: ¿cuántos hay?

            if not eventos:
                messagebox.showinfo("Información",
                                    "No hay registros de eventos para exportar.\n\nUsa 'GUARDAR' para insertar algunos y prueba de nuevo.")
                print("DEBUG EXPORT: 0 eventos - inserta datos en BD")  # Debug
                return

            # Diálogo para elegir dónde guardar
            ruta_archivo = filedialog.asksaveasfilename(
                title="Guardar Excel de Eventos",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile="eventos_exportados.xlsx"  # ← CORREGIDO: initialfile
            )

            if not ruta_archivo:
                print("DEBUG EXPORT: Cancelado por usuario")
                return

            print(f"DEBUG EXPORT: Guardando en {ruta_archivo}")

            # Crear Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Eventos"

            # Headers (para las 11 columnas de eventos)
            headers = ["ID Evento", "ID Recinto", "ID Cliente", "Título", "Tipo", "Fecha Inicio", "Fecha Fin",
                       "Asistentes", "Estado", "Descripción", "Costo","Nombre Recinto","Nombre Cliente",]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")  # Azul
                cell.alignment = Alignment(horizontal="center")

            # Llenar datos (una fila por evento, usando SELECT * que retorna 11 campos)
            for row_idx, evento in enumerate(eventos, 2):  # ← CAMBIO: 'cliente' → 'evento' para claridad
                num_columnas = len(evento)  # ← FIX CLAVE: Número de campos POR EVENTO (ej. 11), no total de eventos
                for col_idx in range(num_columnas):  # ← FIX: range(len(evento)) en lugar de range(11) fijo (dinámico)
                    ws.cell(row=row_idx, column=col_idx + 1,
                            value=evento[col_idx] if col_idx < num_columnas else '')  # ← FIX: Usa len(evento)

            # Ajustar anchos de columnas
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

            # Guardar
            wb.save(ruta_archivo)
            wb.close()

            messagebox.showinfo("Éxito", f"Exportado a:\n{ruta_archivo}\n\nRegistros: {len(eventos)}")
            print(f"DEBUG EXPORT: Éxito con {len(eventos)} registros")

        except Exception as e:
            print(f"DEBUG EXPORT ERROR: {str(e)}")  # Debug
            messagebox.showerror("Error",
                                 f"Error en exportación:\n{str(e)}")  # ← FIX: str(e) completo, sin "si me exporta..."
