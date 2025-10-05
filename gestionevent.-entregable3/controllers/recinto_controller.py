import tkinter as tk
from tkinter import messagebox
from views.recinto_view import RecintoView
from config import COLORES_BOTONES
from tkinter import filedialog  # ← AGREGAR: Para diálogo de guardado
import openpyxl  # ← AGREGAR: pip install openpyxl si no lo tienes
from openpyxl.styles import Font, Alignment, PatternFill  # ← AGREGAR: Para estilos


class RecintoController:
    def __init__(self, parent, app, recinto_model):
        self.app = app
        self.model = recinto_model
        self.view = RecintoView(parent, self)
        self.id_actual = None

    def get_color_boton(self, tipo):
        return COLORES_BOTONES.get(tipo, '#gray')

    def guardar(self):
        valores = self.view.obtener_valores()
        campos_obligatorios = ['nombre', 'ubicacion', 'capacidad', 'tipo', 'tarifa_hora']
        for campo in campos_obligatorios:
            if not valores[campo]:
                messagebox.showerror("Error", f"El campo {campo} es obligatorio")
                return

        try:
            capacidad = int(valores['capacidad'])
            tarifa_hora = float(valores['tarifa_hora'])

            resultado = self.model.guardar(
                valores['nombre'],
                valores.get('foto', ''),
                valores['ubicacion'],
                capacidad,
                valores['tipo'],
                tarifa_hora,
                valores.get('caracteristicas', '')
            )

            if resultado is not None:
                messagebox.showinfo("Éxito", "Recinto guardado correctamente")
                self.view.limpiar_formulario()
                self.app.id_recinto_actual = None

        except ValueError:
            messagebox.showerror("Error", "Capacidad y Tarifa por Hora deben ser valores numéricos")
        except Exception as e:
            messagebox.showerror("Error", f"Error al guardar recinto: {e}")

    def actualizar(self):
        if not self.app.id_recinto_actual:
            messagebox.showwarning("Advertencia", "Primero debe buscar un recinto para actualizar")
            return

        valores = self.view.obtener_valores()
        campos_obligatorios = ['nombre', 'ubicacion', 'capacidad', 'tipo', 'tarifa_hora']
        for campo in campos_obligatorios:
            if not valores[campo]:
                messagebox.showerror("Error", f"El campo {campo} es obligatorio")
                return

        try:
            capacidad = int(valores['capacidad'])
            tarifa_hora = float(valores['tarifa_hora'])

            resultado = self.model.actualizar(
                self.app.id_recinto_actual,
                valores['nombre'],
                valores.get('foto', ''),
                valores['ubicacion'],
                capacidad,
                valores['tipo'],
                tarifa_hora,
                valores.get('caracteristicas', '')
            )

            messagebox.showinfo("Éxito", "Recinto actualizado correctamente")
            self.app.id_recinto_actual = None

        except ValueError:
            messagebox.showerror("Error", "Capacidad y Tarifa por Hora deben ser valores numéricos")
        except Exception as e:
            messagebox.showerror("Error", f"Error al actualizar recinto: {e}")

    def buscar(self):
        ventana, entry, listbox = self.view.ventana_buscar()

        def realizar_busqueda(event=None):
            nombre = entry.get().strip()
            if not nombre:
                return
            self.buscar_recintos(nombre, listbox)

        def seleccionar_recinto(event):
            if not listbox.curselection():
                return
            index = listbox.curselection()[0]
            recinto = listbox.resultados[index]
            self.mostrar_recinto(recinto)
            ventana.destroy()

        entry.bind('<KeyRelease>', realizar_busqueda)
        listbox.bind('<Double-Button-1>', seleccionar_recinto)

    def buscar_recintos(self, nombre, listbox):
        resultados = self.model.buscar_por_nombre(nombre)
        listbox.delete(0, tk.END)
        for resultado in resultados:
            listbox.insert(tk.END, f"{resultado[1]} - {resultado[3]}")  # nombre - ubicacion
        listbox.resultados = resultados

    def mostrar_recinto(self, recinto):
        self.view.mostrar_datos(recinto)
        self.app.id_recinto_actual = recinto[0]

    def eliminar(self):
        if not self.app.id_recinto_actual:
            messagebox.showwarning("Advertencia", "Primero debe buscar un recinto para eliminar")
            return

        if messagebox.askyesno("Confirmar eliminación", "¿Está seguro de que desea eliminar este recinto? Esta acción no se puede deshacer."):
            self.model.eliminar(self.app.id_recinto_actual)
            messagebox.showinfo("Éxito", "Recinto eliminado correctamente")
            self.view.limpiar_formulario()
            self.app.id_recinto_actual = None

    def limpiar(self):
        self.view.limpiar_formulario()
        self.app.id_recinto_actual = None

    def exportar_excel(self):
        """
        Exporta todos los Recintos a un archivo Excel usando ObtenerTodosLosRecintos.
        """
        try:
            print("DEBUG EXPORT: Iniciando exportación...")  # Debug temporal
            recintos = self.model.obtener_todos()  # Llama al método corregido
            print(f"DEBUG EXPORT: Obtenidos {len(recintos)} Servicios de BD")  # Debug: ¿cuántos hay?

            if not recintos:
                messagebox.showinfo("Información",
                                    "No hay registros de recintos para exportar.\n\nUsa 'GUARDAR' para insertar algunos y prueba de nuevo.")
                print("DEBUG EXPORT: 0 recintos - inserta datos en BD")  # Debug
                return

            # Diálogo para elegir dónde guardar
            # Diálogo para elegir dónde guardar
            ruta_archivo = filedialog.asksaveasfilename(
                title="Guardar Excel de recintos",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile="Recintos_exportados.xlsx"  # ← CORREGIDO: Ahora sí, initialfile
            )

            if not ruta_archivo:
                print("DEBUG EXPORT: Cancelado por usuario")
                return

            print(f"DEBUG EXPORT: Guardando en {ruta_archivo}")

            # Crear Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Recintos"

            # Headers (para las 8 columnas típicas de Recintos)
            headers = ["ID Recinto", "Nombre ", "Foto (RUTA)","Ubicacion","Capacidad","Tipo","Tarifa Hora","Caracteristicas"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")  # Azul
                cell.alignment = Alignment(horizontal="center")

            # Llenar datos (una fila por cliente, usando SELECT * que retorna 3 campos)
            for row_idx, recinto in enumerate(recintos, 2):
                num_columnas = len(recinto)
                for col_idx in range(num_columnas):  # Asume 7 columnas
                    ws.cell(row=row_idx, column=col_idx + 1, value=recinto[col_idx] if col_idx < num_columnas else '')

            # Ajustar anchos de columnas
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

            # Guardar
            wb.save(ruta_archivo)
            wb.close()

            messagebox.showinfo("Éxito", f"Exportado a:\n{ruta_archivo}\n\nRegistros: {len(recintos)} recintos en BD")
            print(f"DEBUG EXPORT: Éxito con {len(recintos)} registros")

        except Exception as e:
            print(f"DEBUG EXPORT ERROR: {str(e)}")  # Debug
            messagebox.showerror("Error", f"Error en exportación:\n{str(e)}\n\nRevisa conexión BD y openpyxl.")
