import tkinter as tk
from tkinter import messagebox
from views.servicio_view import ServicioView
from config import COLORES_BOTONES
from tkinter import filedialog  # ← AGREGAR: Para diálogo de guardado
import openpyxl  # ← AGREGAR: pip install openpyxl si no lo tienes
from openpyxl.styles import Font, Alignment, PatternFill  # ← AGREGAR: Para estilos



class ServicioController:
    def __init__(self, parent, app, servicio_model):
        self.app = app
        self.model = servicio_model
        self.view = ServicioView(parent, self)
        self.id_actual = None

    def get_color_boton(self, tipo):
        return COLORES_BOTONES.get(tipo, '#gray')

    def guardar(self):
        valores = self.view.obtener_valores()
        if not valores['nombre'] or not valores['descripcion']:
            messagebox.showerror("Error", "Los campos con * son obligatorios")
            return

        resultado = self.model.guardar(valores['nombre'], valores['descripcion'])
        messagebox.showinfo("Éxito", "Servicio guardado correctamente")
        self.view.limpiar_formulario()
        self.app.id_servicio_actual = None

    def actualizar(self):
        if not self.app.id_servicio_actual:
            messagebox.showwarning("Advertencia", "Primero debe buscar un servicio para actualizar")
            return

        valores = self.view.obtener_valores()
        if not valores['nombre'] or not valores['descripcion']:
            messagebox.showerror("Error", "Los campos con * son obligatorios")
            return

        resultado = self.model.actualizar(self.app.id_servicio_actual, valores['nombre'], valores['descripcion'])
        messagebox.showinfo("Éxito", "Servicio actualizado correctamente")
        self.app.id_servicio_actual = None

    def buscar(self):
        ventana, entry, listbox = self.view.ventana_buscar()

        def realizar_busqueda(event=None):
            nombre = entry.get().strip()
            if not nombre:
                return
            self.buscar_servicios(nombre, listbox)

        def seleccionar_servicio(event):
            if not listbox.curselection():
                return
            index = listbox.curselection()[0]
            servicio = listbox.resultados[index]
            self.mostrar_servicio(servicio)
            ventana.destroy()

        entry.bind('<KeyRelease>', realizar_busqueda)
        listbox.bind('<Double-Button-1>', seleccionar_servicio)

    def buscar_servicios(self, nombre, listbox):
        resultados = self.model.buscar(nombre)
        listbox.delete(0, tk.END)
        for resultado in resultados:
            listbox.insert(tk.END, f"{resultado[1]} - {resultado[2]}")
        listbox.resultados = resultados

    def mostrar_servicio(self, servicio):
        self.view.mostrar_datos(servicio)
        self.app.id_servicio_actual = servicio[0]

    def eliminar(self):
        if not self.app.id_servicio_actual:
            messagebox.showwarning("Advertencia", "Primero debe buscar un servicio para eliminar")
            return

        if messagebox.askyesno("Confirmar eliminación", "¿Está seguro de que desea eliminar este servicio? Esta acción no se puede deshacer."):
            self.model.eliminar(self.app.id_servicio_actual)
            messagebox.showinfo("Éxito", "Servicio eliminado correctamente")
            self.view.limpiar_formulario()
            self.app.id_servicio_actual = None

    def limpiar(self):
        self.view.limpiar_formulario()
        self.app.id_servicio_actual = None

    def exportar_excel(self):
        """
        Exporta todos los Servicios a un archivo Excel usando ObtenerTodosServicios.
        """
        try:
            print("DEBUG EXPORT: Iniciando exportación...")  # Debug temporal
            servicios = self.model.obtener_todos()  # Llama al método corregido
            print(f"DEBUG EXPORT: Obtenidos {len(servicios)} Servicios de BD")  # Debug: ¿cuántos hay?

            if not servicios:
                messagebox.showinfo("Información",
                                    "No hay registros de Servicios para exportar.\n\nUsa 'GUARDAR' para insertar algunos y prueba de nuevo.")
                print("DEBUG EXPORT: 0 servicios- inserta datos en BD")  # Debug
                return

            # Diálogo para elegir dónde guardar
            # Diálogo para elegir dónde guardar
            ruta_archivo = filedialog.asksaveasfilename(
                title="Guardar Excel de servicios",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile="Servicios_exportados.xlsx"  # ← CORREGIDO: Ahora sí, initialfile
            )

            if not ruta_archivo:
                print("DEBUG EXPORT: Cancelado por usuario")
                return

            print(f"DEBUG EXPORT: Guardando en {ruta_archivo}")

            # Crear Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Servicios"

            # Headers (para las 3 columnas típicas de Servicios: id_Servicio, Nombre,Descripcion)
            headers = ["ID Servicio", "Nombre Servicio", "Descripcion"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")  # Azul
                cell.alignment = Alignment(horizontal="center")

            # Llenar datos (una fila por servicio, usando SELECT * que retorna 3 campos)
            for row_idx, servicio in enumerate(servicios, 2):
                num_columnas = len(servicio)
                for col_idx in range(num_columnas):  # Asume 7 columnas
                    ws.cell(row=row_idx, column=col_idx + 1, value=servicio[col_idx] if col_idx < num_columnas else '')

            # Ajustar anchos de columnas
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

            # Guardar
            wb.save(ruta_archivo)
            wb.close()

            messagebox.showinfo("Éxito", f"Exportado a:\n{ruta_archivo}\n\nRegistros: {len(servicios)}")
            print(f"DEBUG EXPORT: Éxito con {len(servicio)} registros")

        except Exception as e:
            print(f"DEBUG EXPORT ERROR: {str(e)}")  # Debug
            messagebox.showerror("Error", f"Error en exportación:\n{str(e)}\n\nRevisa conexión BD y openpyxl.")
