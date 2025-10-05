import tkinter as tk
from tkinter import messagebox
from views.cliente_view import ClienteView
from config import  COLORES_BOTONES
from tkinter import filedialog  # ← AGREGAR: Para diálogo de guardado
import openpyxl  # ← AGREGAR: pip install openpyxl si no lo tienes
from openpyxl.styles import Font, Alignment, PatternFill  # ← AGREGAR: Para estilos



class ClienteController:
    def __init__(self, parent, app, cliente_model):
        self.app = app  # Referencia a la app principal para IDs y colores
        self.model = cliente_model
        self.view = ClienteView(parent, self)
        self.id_actual = None  # Local, pero se sincroniza con app.id_cliente_actual

    def get_color_boton(self, tipo):
        return COLORES_BOTONES.get(tipo, '#gray')

    def guardar(self):
        valores = self.view.obtener_valores()
        if not valores['identificacion'] or not valores['nombre']:
            messagebox.showerror("Error", "Los campos con * son obligatorios")
            return

        resultado = self.model.guardar(
            valores['identificacion'],
            valores.get('foto', ''),
            valores['nombre'],
            valores.get('telefono', ''),
            valores.get('email', ''),
            valores.get('direccion', '')
        )

        if resultado is not None:
            messagebox.showinfo("Éxito", "Cliente guardado correctamente")
            self.view.limpiar_formulario()
            self.app.id_cliente_actual = None

    def actualizar(self):
        if not self.app.id_cliente_actual:
            messagebox.showwarning("Advertencia", "Primero debe buscar un cliente para actualizar")
            return

        valores = self.view.obtener_valores()
        if not valores['identificacion'] or not valores['nombre']:
            messagebox.showerror("Error", "Los campos con * son obligatorios")
            return

        resultado = self.model.actualizar(
            self.app.id_cliente_actual,
            valores['identificacion'],
            valores.get('foto', ''),
            valores['nombre'],
            valores.get('telefono', ''),
            valores.get('email', ''),
            valores.get('direccion', '')
        )

        if resultado:
            messagebox.showinfo("Éxito", "Cliente actualizado correctamente")
            self.app.id_cliente_actual = None  # Reset después de actualizar

    def buscar(self):
        ventana, entry = self.view.ventana_buscar()

        def realizar_busqueda():
            identificacion = entry.get().strip()
            if not identificacion:
                messagebox.showwarning("Advertencia", "Por favor ingrese una identificación")
                return

            resultado = self.model.buscar_por_identificacion(identificacion)
            if resultado:
                cliente = resultado[0]
                self.mostrar_cliente(cliente)
                self.app.id_cliente_actual = cliente[0]
                ventana.destroy()
            else:
                messagebox.showinfo("Información", "No se encontró ningún cliente con esa identificación")

        btn_buscar = tk.Button(ventana, text="Buscar", command=realizar_busqueda, bg='#3498db', fg='white')
        btn_buscar.pack(pady=10)

    def mostrar_cliente(self, cliente):
        self.view.mostrar_datos(cliente)
        self.id_actual = cliente[0]

    def eliminar(self):
        if not self.app.id_cliente_actual:
            messagebox.showwarning("Advertencia", "Primero debe buscar un cliente para eliminar")
            return

        if messagebox.askyesno("Confirmar eliminación", "¿Está seguro de que desea eliminar este cliente? Esta acción no se puede deshacer."):
            resultado = self.model.eliminar(self.app.id_cliente_actual)
            if resultado:
                messagebox.showinfo("Éxito", "Cliente eliminado correctamente")
                self.view.limpiar_formulario()
                self.app.id_cliente_actual = None

    def limpiar(self):
        self.view.limpiar_formulario()
        self.app.id_cliente_actual = None

    def exportar_excel(self):
        """
        Exporta todos los clientes a un archivo Excel usando ObtenerTodosClientes.
        """
        try:
            print("DEBUG EXPORT: Iniciando exportación...")  # Debug temporal
            clientes = self.model.obtener_todos()  # Llama al método corregido
            print(f"DEBUG EXPORT: Obtenidos {len(clientes)} clientes de BD")  # Debug: ¿cuántos hay?

            if not clientes:
                messagebox.showinfo("Información",
                                    "No hay registros de clientes para exportar.\n\nUsa 'GUARDAR' para insertar algunos y prueba de nuevo.")
                print("DEBUG EXPORT: 0 clientes - inserta datos en BD")  # Debug
                return

            # Diálogo para elegir dónde guardar
            # Diálogo para elegir dónde guardar
            ruta_archivo = filedialog.asksaveasfilename(
                title="Guardar Excel de Clientes",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile="clientes_exportados.xlsx"  # ← CORREGIDO: Ahora sí, initialfile
            )

            if not ruta_archivo:
                print("DEBUG EXPORT: Cancelado por usuario")
                return

            print(f"DEBUG EXPORT: Guardando en {ruta_archivo}")

            # Crear Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Clientes"

            # Headers (para las 7 columnas típicas de clientes: id, identificacion, foto, nombre, telefono, email, direccion)
            headers = ["ID Cliente", "Identificación", "Foto (Ruta)", "Nombre", "Teléfono", "Email", "Dirección"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")  # Azul
                cell.alignment = Alignment(horizontal="center")

            # Llenar datos (una fila por cliente, usando SELECT * que retorna 7 campos)
            for row_idx, cliente in enumerate(clientes, 2):
                num_columnas = len(cliente)
                for col_idx in range(num_columnas):  # Asume 7 columnas
                    ws.cell(row=row_idx, column=col_idx + 1, value=cliente[col_idx] if col_idx < num_columnas else '')

            # Ajustar anchos de columnas
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

            # Guardar
            wb.save(ruta_archivo)
            wb.close()

            messagebox.showinfo("Éxito", f"Exportado a:\n{ruta_archivo}\n\nRegistros: {len(clientes)}")
            print(f"DEBUG EXPORT: Éxito con {len(clientes)} registros")

        except Exception as e:
            print(f"DEBUG EXPORT ERROR: {str(e)}")  # Debug
            messagebox.showerror("Error", f"Error en exportación:\n{str(e)}\n\nRevisa conexión BD y openpyxl.")